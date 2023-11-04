import openpyxl
from django.shortcuts import render,redirect
from django.shortcuts import render, get_object_or_404
# from django.http import JsonResponse
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from .models import Client, Trade
from .forms import AddClientForm
from django.http import HttpResponse

# Create your views here.
    
def home(request):
    if request.method == 'POST':
        status = request.POST.get('status')
        client_id1 = request.POST.get('client_id')
        name = request.POST.get('name')
        exchange = request.POST.get('exchange')
        
        # if not status or not client_id1 or not name or not exchange:
        #     error_message = "All fields are required."
        #     # You can customize the error handling as needed, e.g., by rendering the error message
        #     return render(request, 'home.html', {'error_message': error_message})
        
        try:
            client = Client.objects.get(client_id=client_id1)
        except Client.DoesNotExist:
            client = None

        if client:
        # Fetch the brokerage rate for the client from the Client model
            brokerage_rate = client.brokerage
        else:
            brokerage_rate = 0.06  # Set a default brokerage rate if client is not found
    # print(brokerage_rate)
        
        if exchange == 'mcx':
            symbol = request.POST.get('mcx_symbol')
            lot_size = request.POST.get('lot_size')
            # Check if lot_size is not empty before converting to int
            qty = int(lot_size) if lot_size else 0  # Set to 0 if empty
        elif exchange == 'nse':
            symbol = request.POST.get('nse_symbol')
            qty_size = request.POST.get('qty_size')
            qty = int(qty_size) if qty_size else 0  # Set to 0 if empty
            
        rate = float(request.POST.get('rate'))
        
        if status == 'buy':
            netrate = float(((rate * brokerage_rate)/100) + rate)
            
        elif status == 'sell':
            netrate = float( rate - ((rate * brokerage_rate)/100))
        
        amount = netrate * qty
        
        # if (status =="" or client == "" or name == "" or exchange == "" or symbol =="" or qty == "" or rate == "" or netrate == "" or amount == ""):
        #     print("error") 
            
        # else:
        #     pass
        
        try:
            client = Client.objects.get(client_id=client_id1)
            print('helloo',client)
        except Client.DoesNotExist:
            pass
        
        # print(status,client,name,exchange,symbol,qty,rate,netrate,amount)
        trade = Trade(status=status,client_id=client,name=name,exchange=exchange,symbol=symbol,qty=qty,rate=rate,netrate=netrate,amount=amount)
        trade.save()
        return redirect('home') 
    
    clnt_data = reversed(Trade.objects.all())
    client_list = Client.objects.all()
    
    return render(request,'home.html',{'clnt_data':clnt_data,'client_list':client_list})

def show(request):
    clnt_data = reversed(Client.objects.all())
    return render(request,'show.html',{'clnt_data':clnt_data})

def client_trade(request, client_id):
    client = get_object_or_404(Client, client_id=client_id)
    trades = Trade.objects.filter(client_id=client)
    
    buy_trades = Trade.objects.filter(client_id=client, status='buy')
    sell_trades = Trade.objects.filter(client_id=client, status='sell')

    total_buy_amount = sum([trade.amount for trade in buy_trades])
    # print(total_buy_amount)
    total_sell_amount = sum([trade.amount for trade in sell_trades])
    
    final_amount =  total_sell_amount - total_buy_amount

    # context = {
    #     'client': client,
    #     'buy_trades': buy_trades,
    #     'sell_trades': sell_trades,
    #     'total_buy_amount': total_buy_amount,
    #     'total_sell_amount': total_sell_amount,
    # }

    return render(request, 'client_trade.html', {'client': client, 'trades': trades,'total_buy_amount': total_buy_amount,'total_sell_amount': total_sell_amount,'final_amount':final_amount,'client_id': client_id})

def add(request):
    if request.method == 'POST':
     fm =AddClientForm(request.POST)
     if fm.is_valid():
        no = fm.cleaned_data['client_id']
        nm = fm.cleaned_data['name']
        br = fm.cleaned_data['brokerage']
        dt = fm.cleaned_data['date']
        reg = Client(client_id=no, name=nm, brokerage=br, date=dt)
        reg.save()
        fm = AddClientForm()
        return redirect('show')
    else:
     fm = AddClientForm()
    return render(request,"add.html",{'form':fm})

def edit(request,client_id):
    if request.method == 'POST':
      pi = Client.objects.get(pk=client_id)
      fm = AddClientForm(request.POST, instance=pi)
      if fm.is_valid():
         fm.save()
         return redirect('show')
    else:
       pi = Client.objects.get(pk=client_id)
       fm = AddClientForm(instance=pi)
    return render(request,'edit.html',{'form':fm})

def delete(request,id):
   if request.method == 'POST':
      pi = Client.objects.get(pk=id)
      pi.delete()
      return redirect('show')

def export_trade_history_to_excel(request, client_id):
    # Fetch all trades for the specified client
    trades = Trade.objects.filter(client_id=client_id)
    
    total_buy_amount = sum(trade.amount for trade in trades if trade.status == 'buy')
    total_sell_amount = sum(trade.amount for trade in trades if trade.status == 'sell')
    final_amount = total_sell_amount - total_buy_amount

    # Create a new workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trade History"
    

    # Define column headers and set column widths
    columns = ["Index", "Client ID","Status", "Symbol", "Qty", "Rate", "NetRate", "Amount"]
    for col_num, column_title in enumerate(columns, 1):
        col_letter = get_column_letter(col_num)
        cell = ws["{}1".format(col_letter)]
        cell.value = column_title
        cell.font = Font(bold=True)
        ws.column_dimensions[col_letter].width = 15

    row_colors = {
        "buy": PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid"),  # Green
        "sell": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),  # Red
    }
    
    # Populate the worksheet with trade data
    for index, trade in enumerate(trades, 2):
        ws.cell(row=index, column=1, value=index - 1)  # Index
        ws.cell(row=index, column=2, value=trade.client_id.client_id)  # Client ID
        ws.cell(row=index, column=3, value=trade.status)  # Status
        ws.cell(row=index, column=4, value=trade.symbol)  # Symbol
        ws.cell(row=index, column=5, value=trade.qty)  # Qty
        ws.cell(row=index, column=6, value=trade.rate)  # Rate
        ws.cell(row=index, column=7, value=trade.netrate)  # NetRate
        ws.cell(row=index, column=8, value=trade.amount)  # Amount
        
        # Apply background color based on status
        if trade.status in row_colors:
            for col_num in range(1, len(columns) + 1):
                col_letter = get_column_letter(col_num)
                ws["{}{}".format(col_letter, index)].fill = row_colors[trade.status]
    

        ws.cell(row=len(trades) + 4, column=7, value="Total Buy Amount:")
        ws.cell(row=len(trades) + 4, column=8, value=total_buy_amount)
        ws.cell(row=len(trades) + 5, column=7, value="Total Sell Amount:")
        ws.cell(row=len(trades) + 5, column=8, value=total_sell_amount)
        ws.cell(row=len(trades) + 6, column=7, value="Final Amount:")
        ws.cell(row=len(trades) + 6, column=8, value=final_amount)

    # Create an HTTP response with the Excel file
    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = f'attachment; filename="trade_history_client_{client_id}.xlsx"'
    wb.save(response)

    return response
  